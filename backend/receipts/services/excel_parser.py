"""
Excel Parser Service - Validates and parses Excel files for receipt upload.
"""
import logging
from decimal import Decimal, InvalidOperation
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger('receipts')


class ExcelParserService:
    """
    Service for parsing and validating Excel files containing receipt data.
    
    Expected Excel columns:
    - receipt_number (required, unique)
    - student_name (required)
    - class_name (required)
    - payment_mode (required: cash/cheque/bank_transfer/upi/card/other)
    - date (required, format: YYYY-MM-DD or DD/MM/YYYY)
    - annual_fee (optional, numeric)
    - tuition_fee (optional, numeric)
    - kit_books_fee (optional, numeric)
    - activity_fee (optional, numeric)
    - uniform_fee (optional, numeric)
    """
    
    REQUIRED_HEADERS = [
        'receipt_number',
        'student_name', 
        'class_name',
        'payment_mode',
        'date',
    ]
    
    OPTIONAL_HEADERS = [
        'annual_fee',
        'tuition_fee',
        'kit_books_fee',
        'activity_fee',
        'uniform_fee',
    ]
    
    ALL_HEADERS = REQUIRED_HEADERS + OPTIONAL_HEADERS
    
    VALID_PAYMENT_MODES = [
        'cash',
        'cheque',
        'bank_transfer',
        'upi',
        'card',
        'other',
    ]
    
    # Column name aliases for flexibility
    HEADER_ALIASES = {
        'receipt_number': ['receipt_no', 'receiptnumber', 'receipt no', 'receiptno'],
        'student_name': ['student', 'name', 'studentname', 'student name'],
        'class_name': ['class', 'classname', 'class name', 'grade'],
        'payment_mode': ['payment', 'paymentmode', 'payment mode', 'mode'],
        'date': ['receipt_date', 'receiptdate', 'payment_date'],
        'annual_fee': ['annual', 'annualfee', 'annual fee'],
        'tuition_fee': ['tuition', 'tuitionfee', 'tuition fee'],
        'kit_books_fee': ['kit_books', 'kitbooks', 'kit books', 'kitbooksfee', 'kit books fee'],
        'activity_fee': ['activity', 'activityfee', 'activity fee'],
        'uniform_fee': ['uniform', 'uniformfee', 'uniform fee'],
    }
    
    def __init__(self, file_path: str = None, file_object=None):
        """
        Initialize parser with file path or file object.
        
        Args:
            file_path: Path to Excel file
            file_object: Django UploadedFile object
        """
        self.file_path = file_path
        self.file_object = file_object
        self.workbook: Optional[Workbook] = None
        self.worksheet: Optional[Worksheet] = None
        self.header_mapping: Dict[str, int] = {}
        self.errors: List[Dict[str, Any]] = []
        self.warnings: List[str] = []
        
    def load_workbook(self) -> Tuple[bool, str]:
        """
        Load the Excel workbook.
        
        Returns:
            Tuple of (success, message)
        """
        try:
            if self.file_object:
                self.workbook = openpyxl.load_workbook(self.file_object)
            elif self.file_path:
                self.workbook = openpyxl.load_workbook(self.file_path)
            else:
                return False, "No file provided"
            
            # Get first worksheet
            self.worksheet = self.workbook.active
            return True, "Workbook loaded successfully"
            
        except Exception as e:
            logger.error(f"Failed to load Excel file: {str(e)}")
            return False, f"Failed to load Excel file: {str(e)}"
    
    def normalize_header(self, header: str) -> Optional[str]:
        """
        Normalize header name to match expected headers.
        
        Args:
            header: Raw header string from Excel
            
        Returns:
            Normalized header name or None if not recognized
        """
        if not header:
            return None
            
        header_lower = str(header).strip().lower().replace(' ', '_').replace('-', '_')
        
        # Direct match
        if header_lower in self.ALL_HEADERS:
            return header_lower
            
        # Check aliases
        for standard_name, aliases in self.HEADER_ALIASES.items():
            if header_lower in [a.lower().replace(' ', '_') for a in aliases]:
                return standard_name
                
        return None
    
    def validate_headers(self) -> Tuple[bool, List[str]]:
        """
        Validate that all required headers are present.
        
        Returns:
            Tuple of (is_valid, missing_headers)
        """
        if not self.worksheet:
            return False, ["Worksheet not loaded"]
        
        # Get first row as headers
        header_row = list(self.worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        self.header_mapping = {}
        missing_headers = []
        
        for col_idx, header in enumerate(header_row):
            normalized = self.normalize_header(header)
            if normalized:
                self.header_mapping[normalized] = col_idx
        
        # Check required headers
        for required in self.REQUIRED_HEADERS:
            if required not in self.header_mapping:
                missing_headers.append(required)
        
        is_valid = len(missing_headers) == 0
        
        if not is_valid:
            logger.warning(f"Missing required headers: {missing_headers}")
        
        return is_valid, missing_headers
    
    def parse_date(self, value: Any) -> Optional[datetime]:
        """
        Parse date from various formats.
        
        Args:
            value: Date value from Excel
            
        Returns:
            Parsed date or None
        """
        if value is None:
            return None
            
        # If already a datetime object
        if isinstance(value, datetime):
            return value.date()
        
        # If it's a string, try various formats
        if isinstance(value, str):
            date_formats = [
                '%Y-%m-%d',
                '%d/%m/%Y',
                '%d-%m-%Y',
                '%m/%d/%Y',
                '%d.%m.%Y',
            ]
            for fmt in date_formats:
                try:
                    return datetime.strptime(value.strip(), fmt).date()
                except ValueError:
                    continue
        
        # If it's a number (Excel serial date)
        if isinstance(value, (int, float)):
            try:
                from openpyxl.utils import get_column_letter
                from datetime import timedelta
                # Excel serial date (days since 1899-12-30)
                return datetime(1899, 12, 30) + timedelta(days=int(value))
            except:
                pass
                
        return None
    
    def parse_decimal(self, value: Any) -> Decimal:
        """
        Parse decimal value from Excel.
        
        Args:
            value: Numeric value from Excel
            
        Returns:
            Decimal value (0 if invalid)
        """
        if value is None:
            return Decimal('0.00')
            
        try:
            if isinstance(value, (int, float)):
                return Decimal(str(value)).quantize(Decimal('0.01'))
            elif isinstance(value, str):
                # Remove currency symbols and commas
                cleaned = value.replace(',', '').replace('₹', '').replace('$', '').strip()
                return Decimal(cleaned).quantize(Decimal('0.01'))
        except (InvalidOperation, ValueError):
            pass
            
        return Decimal('0.00')
    
    def parse_payment_mode(self, value: Any) -> Optional[str]:
        """
        Parse and validate payment mode.
        
        Args:
            value: Payment mode string
            
        Returns:
            Normalized payment mode or None
        """
        if not value:
            return None
            
        mode_lower = str(value).strip().lower().replace(' ', '_').replace('-', '_')
        
        # Direct match
        if mode_lower in self.VALID_PAYMENT_MODES:
            return mode_lower
            
        # Common variations
        mode_mapping = {
            'cash': 'cash',
            'cheque': 'cheque',
            'check': 'cheque',
            'bank': 'bank_transfer',
            'bank_transfer': 'bank_transfer',
            'neft': 'bank_transfer',
            'rtgs': 'bank_transfer',
            'imps': 'bank_transfer',
            'upi': 'upi',
            'gpay': 'upi',
            'googlepay': 'upi',
            'phonepe': 'upi',
            'card': 'card',
            'credit_card': 'card',
            'debit_card': 'card',
            'credit': 'card',
            'debit': 'card',
            'other': 'other',
            'others': 'other',
        }
        
        return mode_mapping.get(mode_lower)
    
    def parse_row(self, row: Tuple, row_number: int) -> Tuple[bool, Dict[str, Any]]:
        """
        Parse a single row of Excel data.
        
        Args:
            row: Tuple of cell values
            row_number: Row number for error reporting
            
        Returns:
            Tuple of (is_valid, parsed_data)
        """
        data = {}
        errors = []
        
        # Extract values using header mapping
        def get_value(header_name: str) -> Any:
            col_idx = self.header_mapping.get(header_name)
            if col_idx is not None and col_idx < len(row):
                return row[col_idx]
            return None
        
        # Required fields
        receipt_number = get_value('receipt_number')
        if not receipt_number:
            errors.append(f"Row {row_number}: Missing receipt_number")
        else:
            data['receipt_number'] = str(receipt_number).strip()
        
        student_name = get_value('student_name')
        if not student_name:
            errors.append(f"Row {row_number}: Missing student_name")
        else:
            data['student_name'] = str(student_name).strip()
        
        class_name = get_value('class_name')
        if not class_name:
            errors.append(f"Row {row_number}: Missing class_name")
        else:
            data['class_name'] = str(class_name).strip()
        
        payment_mode = self.parse_payment_mode(get_value('payment_mode'))
        if not payment_mode:
            errors.append(f"Row {row_number}: Invalid or missing payment_mode")
        else:
            data['payment_mode'] = payment_mode
        
        date = self.parse_date(get_value('date'))
        if not date:
            errors.append(f"Row {row_number}: Invalid or missing date")
        else:
            data['date'] = date
        
        # Optional fee fields
        data['annual_fee'] = self.parse_decimal(get_value('annual_fee'))
        data['tuition_fee'] = self.parse_decimal(get_value('tuition_fee'))
        data['kit_books_fee'] = self.parse_decimal(get_value('kit_books_fee'))
        data['activity_fee'] = self.parse_decimal(get_value('activity_fee'))
        data['uniform_fee'] = self.parse_decimal(get_value('uniform_fee'))
        
        # Calculate total
        data['total_amount'] = (
            data['annual_fee'] +
            data['tuition_fee'] +
            data['kit_books_fee'] +
            data['activity_fee'] +
            data['uniform_fee']
        )
        
        is_valid = len(errors) == 0
        
        if errors:
            for error in errors:
                self.errors.append({'row': row_number, 'error': error})
        
        return is_valid, data
    
    def parse(self) -> Tuple[bool, List[Dict[str, Any]], Dict[str, Any]]:
        """
        Parse the entire Excel file.
        
        Returns:
            Tuple of (success, parsed_rows, summary)
        """
        # Load workbook
        success, message = self.load_workbook()
        if not success:
            return False, [], {'error': message}
        
        # Validate headers
        valid, missing = self.validate_headers()
        if not valid:
            return False, [], {
                'error': f"Missing required headers: {', '.join(missing)}",
                'missing_headers': missing
            }
        
        parsed_rows = []
        row_count = 0
        
        # Skip header row, iterate through data rows
        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if all(cell is None for cell in row):
                continue
            
            row_count += 1
            is_valid, data = self.parse_row(row, row_idx)
            
            if is_valid:
                parsed_rows.append(data)
        
        summary = {
            'total_rows': row_count,
            'valid_rows': len(parsed_rows),
            'invalid_rows': row_count - len(parsed_rows),
            'error_count': len(self.errors),
            'errors': self.errors[:50],  # Limit errors returned
        }
        
        success = len(parsed_rows) > 0
        
        logger.info(f"Excel parsing complete: {len(parsed_rows)} valid rows out of {row_count}")
        
        return success, parsed_rows, summary
    
    def close(self):
        """Close the workbook to free resources."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.worksheet = None